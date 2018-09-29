def test_after_150_days() :
    import os #for giving adress to where we want to save output files
    from openpyxl import load_workbook
    wb = load_workbook('TestDataset_after_150_days.xlsx')
    ws = wb.active
    name=[]
    number_of_posts = []
    number_of_followers = []
    number_of_followings = []
    number_of_likes_for_10th_ex_post = []
    number_of_likes_for_11th_ex_post = []
    number_of_likes_for_12th_ex_post = []
    number_of_self_picture_posts_form_9_previous_posts = []
    sex = []
    
    
    #getting data of each column of excel file into a list named by the name of corrospondant name of that column 
    for row in ws :
        if row[2].value != None :  # because after 150 days the profile of some users were not availbale so we need to delete their rows by this line (look at the dataset)
            name.append(row[0].value)
            number_of_posts.append(row[1].value)
            number_of_followers.append(row[2].value)
            number_of_followings.append(row[3].value)
            number_of_likes_for_10th_ex_post.append(row[4].value)
            number_of_likes_for_11th_ex_post.append(row[5].value)
            number_of_likes_for_12th_ex_post.append(row[6].value)
            number_of_self_picture_posts_form_9_previous_posts.append(row[7].value)
            sex.append(row[8].value)

#------------------------------------
    #finding the number of users in dataset
    number_of_users_in_dataset = len (name)    
    print ("number_of_users_in_dataset",number_of_users_in_dataset)

#------------------------------------        
    #finding the number of male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    f_number = 0
    m_number = 0
    while i < number_of_users_in_dataset : #because total number of users is 1000 and the first row is for labels so we have to go up to 1001
        if sex[i] == "f" :
            f_number += 1

        elif sex[i] == "m" :
            m_number += 1
        else :
            print("aha")


        i+=1

#------------------------------------
    #finding the mean number of self posts (from 9 posts) for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_self_post_for_all_females = 0
    sum_self_post_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_self_picture_posts_form_9_previous_posts[i]
            sum_self_post_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_self_picture_posts_form_9_previous_posts[i]
            sum_self_post_for_all_males += int(w)
            
        i+=1
        
    mean_of_self_post_for_females = sum_self_post_for_all_females/f_number
    mean_of_self_post_for_males = sum_self_post_for_all_males/m_number

#-----------------------------------
    #finding the mean number of followers for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_followers_for_all_females = 0
    sum_number_of_followers_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_followers[i]
            sum_number_of_followers_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_followers[i]
            sum_number_of_followers_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_followers_for_females = sum_number_of_followers_for_all_females/f_number
    mean_number_of_followers_for_males = sum_number_of_followers_for_all_males/m_number
    
#-------------------------------------
    #finding the mean number of followings for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_followings_for_all_females = 0
    sum_number_of_followings_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_followings[i]
            sum_number_of_followings_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_followings[i]
            sum_number_of_followings_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_followings_for_females = sum_number_of_followings_for_all_females/f_number
    mean_number_of_followings_for_males = sum_number_of_followings_for_all_males/m_number

#-------------------------------------
    #finding the mean number of posts for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_posts_for_all_females = 0
    sum_number_of_posts_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_posts[i]
            sum_number_of_posts_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_posts[i]
            sum_number_of_posts_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_posts_for_females = sum_number_of_posts_for_all_females/f_number
    mean_number_of_posts_for_males = sum_number_of_posts_for_all_males/m_number
#-------------------------------------
    #finding the relation between number of self posts(in 9 previous posts) and following/follower 
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    ratio_of_followings_to_followers_for_all_users = []
    while i < number_of_users_in_dataset : 
        t = number_of_followings[i]/number_of_followers[i]
        ratio_of_followings_to_followers_for_all_users.append(t)   
        i+=1
    i = 0

    del number_of_self_picture_posts_form_9_previous_posts[0] #before this command, "number_of_self_picture_posts_form_9_previous_posts" was a list by 1001 elemnts, remember the first element was the label, but we need a that list to just have 1000 elemnts, so be the same size of "ratio_of_followings_to_followers_for_all_users". by having both a 1000 elements list then we can compare and analyse them.

#---------------------------------------------------------
   #finding the relation between the mean number of likes for posts(based on ex 10th 11th 12th posts) and the ratio of following/follower
    i=1
    mean_like_for_all_users = []
    while i < number_of_users_in_dataset :
        total_like = number_of_likes_for_10th_ex_post[i] + number_of_likes_for_11th_ex_post[i] + number_of_likes_for_12th_ex_post[i]
        mean_like = total_like/3
        mean_like_for_all_users.append (mean_like)
        i+=1
##    import matplotlib.pyplot as plt
##    import numpy
##    plt.hist2d(mean_like_for_all_users,ratio_of_followings_to_followers_for_all_users,50)
##    plt.colorbar()
##    plt.show()

#-------- Hist2D : like/following & following/follower
    #finding the relation between the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/following" and the ratio of "following/follower"
    i=0
    mean_like_to_following_for_all_users = []
    del number_of_followings[0]
    while i < number_of_users_in_dataset -1 :
        mean_like_to_following = mean_like_for_all_users[i]/number_of_followings[i]
        mean_like_to_following_for_all_users.append (mean_like_to_following)
        i+=1
    import matplotlib.pyplot as plt
    import numpy
#####################################    print("MAX[mean_like_to_following_for_all_users] = " , max(mean_like_to_following_for_all_users)) #we need this line to set max limit for horizental axis in plt.hist2d command (two lines below), which is 32.12 (we set to 33 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
#####################################    print("MAX[ratio_of_followings_to_followers_for_all_users] = " , max(ratio_of_followings_to_followers_for_all_users)) #we need this line to set max limit for vertical axis in plt.hist2d command (two lines below), which is 15.66 (we set to 16 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
    plt.hist2d(mean_like_to_following_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,2],[0,16]]) #notice max horizental was 33 but since there was only a few data with more than 2 as ratio of like to following so we ignored them and set the max horizental limit to 2
##    plt.scatter(mean_like_to_following_for_all_users,ratio_of_followings_to_followers_for_all_users)
    plt.xlabel('Like/Following ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Following & Following/Follower',fontsize=12)
    file_name = "Hist2D_LikeFollowing_to_FollowingFollower(after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
#-------- Hist2D : like/follower & following/follower
    #finding the relation between the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/follower" and the ratio of "following/follower"
    i=0
    mean_like_to_follower_for_all_users = []
    del number_of_followers[0]
    while i < number_of_users_in_dataset - 1 :
        mean_like_to_follower = mean_like_for_all_users[i]/number_of_followers[i]
        mean_like_to_follower_for_all_users.append (mean_like_to_follower)
        i+=1
    import matplotlib.pyplot as plt
    import numpy as np
    import matplotlib.colors as mcolors
    import matplotlib.mlab as mlab
    import os #To give address for saving output plots
##    import scipy.stats

##    from numpy._distributor_init import NUMPY_MKL 
##    from scipy.stats import norm
#####################################    print("MAX[mean_like_to_follower_for_all_users] = " , max(mean_like_to_follower_for_all_users)) #we need this line to set max limit for horizental axis in plt.hist2d command (two lines below), which is 0.92 (we set to 1 ) , notice once we run the program and see the 0.92 then we set it in plt.hist2d for ever
#####################################    print("MAX[ratio_of_followings_to_followers_for_all_users] = " , max(ratio_of_followings_to_followers_for_all_users)) #we need this line to set max limit for vertical axis in plt.hist2d command (two lines below), which is 15.66 (we set to 16 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
    plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,1],[0,16]])
    
##    counts, _ , _ , _ = plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,1],[0,16]])
##    row_total = 0
##    total = 0
##    for w in counts :
##        row_total += w
##    for y in row_total :
##        total += y
##    print(total)
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Follower & Following/Follower',fontsize=12)
    file_name = "Hist2D_LikeFollower_to_FollowingFollower(after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    v = [0 , 1 , 2 , 3 , 4 , 5 , 6 , 7 , 8 ]
    cbar = plt.colorbar(ticks = v) #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #---- Hist2D : like/follower & following/follower --- its the above hist2d but here we just focus on the dense area  
    plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,33,[[0,0.5],[0,10]])
##    plt.scatter(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users)
##    plt.show()
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Follower & Following/Follower (focused)',fontsize=12)
    file_name = "Hist2Dfocused_LikeFollower_to_FollowingFollower(after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #-------- Hist1D : following/follower 

    plt.hist(ratio_of_followings_to_followers_for_all_users,[0,0.5,1,1.5,2,2.5,3,3.5,4,4.5,5,5.5,6,6.5,7,7.5,8,8.5,9,9.5,10,10.5,11,11.5,12,12.5,13,13.5,14,14.5,15,15.5,16]) #we set the edges of bins for histogram from 0 to 16(which is the max in the dataset and we noticed that from previous command  
    plt.ylim(0,200)
    #-----
##    pdf_x = np.linspace(np.min(ratio_of_followings_to_followers_for_all_users),np.max(ratio_of_followings_to_followers_for_all_users),100)
##    pdf_y = 1.0/np.sqrt(2*np.pi*var)*np.exp(-0.5*(pdf_x-avg)**2/var)
##    plt.plot(pdf_x,pdf_y,'k--')
##    print(avg)
##    print(var)
##    plt.colorbar()
##    (mu, sigma) = norm.fit(datos)
##    y = mlab.normpdf( bins, mu, sigma)
##    l = plt.plot(bins, y, 'r--', linewidth=2)
     
    plt.xlabel('Following/Follower ratio')
    plt.ylabel('Frequency')
##    plt.title('Following/Follower',fontsize=12)
    file_name = "Hist1D_Following_to_Follower (after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #-------- Hist1D : like/follower

    plt.hist(mean_like_to_follower_for_all_users,[0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]) #we set the edges of bins for histogram from 0 to 0.92(which is the max in the dataset and we noticed that from previous commands  
    #--------
##    pdf_x = np.linspace(np.min(ratio_of_followings_to_followers_for_all_users),np.max(ratio_of_followings_to_followers_for_all_users),100)
##    pdf_y = 1.0/np.sqrt(2*np.pi*var)*np.exp(-0.5*(pdf_x-avg)**2/var)
##    plt.plot(pdf_x,pdf_y,'k--')
##    print(avg)
##    print(var)
##    plt.colorbar()
##    (mu, sigma) = norm.fit(datos)
##    y = mlab.normpdf( bins, mu, sigma)
##    l = plt.plot(bins, y, 'r--', linewidth=2)
     
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Frequency')
##    plt.title('Like/Follower',fontsize=12)
    file_name = "Hist1D_Like_to_Follower (after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #-------- 

    

#-----------------------------------------------------------
    #great result :
    #THOSE USERS WITH FOLLOWEING/FOLLOWER < 1 WITH STRONGER PROBABILITY HAVE BETTER CHANCE OF ACCEPTANCE . In other words they have better like/follower (like/following) , so the chance that they start the rumor are more stronger 
#-----------------------------------------------------------    
    #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/follower"
    #PLEASE notice that if we use discrete values of 0,1,2,3,..,9 as number of self posts(narcissism) then hist2d shape would be discrete that is not good, so for each number we randomly choose a value between two defined limit. for example if self posts number is 2 then we choose ranmoly a number between (0.2 , 0.3) to represent its narcissism level
    i=0
    import matplotlib.pyplot as plt
    import numpy
    import random
    narcissism = []
    while i < number_of_users_in_dataset - 1 :
        if number_of_self_picture_posts_form_9_previous_posts[i] == 0 :
            t = random.uniform(0,0.1)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 1 :
            t = random.uniform(0.1,0.2)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 2 :
            t = random.uniform(0.2,0.3)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 3 :
            t = random.uniform(0.3,0.4)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 4 :
            t = random.uniform(0.4,0.5)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 5 :
            t = random.uniform(0.5,0.6)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 6 :
            t = random.uniform(0.6,0.7)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 7 :
            t = random.uniform(0.7,0.8)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 8 :
            t = random.uniform(0.8,0.9)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 9 :
            t = random.uniform(0.9,1)
            narcissism.append (t)
        i+=1
    plt.hist2d(narcissism,mean_like_to_follower_for_all_users,8,[[0,1],[0,1]])
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('Like/Follower')
##    plt.title('narcissism & Like/Follower ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_LikeFollower (after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !

    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
#---- #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the ratio of following/follower"
    plt.hist2d(narcissism,ratio_of_followings_to_followers_for_all_users,9,[[0,1],[0,16]])
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('Following/Follower')
##    plt.title('narcissism & Following/Follower ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_FollowingFollower (after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
##----------
    #---- #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the number of posts
    i=1
    del number_of_posts[0]
    plt.hist2d(narcissism,number_of_posts,40,[[0,1],[50,2000]]) #notice the min of posts was 50 because we didnt pick users with less than 50 posts and also max was 10000 but since there was a few users with more than 2000 so we made up limit to 2000
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('number_of_posts')
##    plt.title('narcissism & number_of_posts ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_numberofposts (after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #-----
    #---- #finding the following/follower ratio and the number of posts
    i=1
    plt.hist2d(ratio_of_followings_to_followers_for_all_users,number_of_posts,40,[[0,16],[50,2000]]) #notice the min of posts was 50 because we didnt pick users with less than 50 posts and also max was 10000 but since there was a few users with more than 2000 so we made up limit to 2000
    plt.xlabel('following/follower')
    plt.ylabel('number_of_posts')
    plt.title('following/follower & number_of_posts ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_numberofposts (after 150 days)"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_After_150_Days_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()

    

# finding the relation of self post with  "following/follower" , "like/follower" and "sex"
# notic we want to find best amount for cc , cx , cxx , cy , cyy , cq in the equation z = cc + cx*x + cxx*x^2 + cxxx*x^3 + cy*y + cyy*y^2 + cyyy*y^3 + cq*q (in which x , y , q are independent variables)
    import matplotlib.pyplot as plt
    import numpy as np
    i = 1
    sex_values = []
    while i < number_of_users_in_dataset :
        if sex[i] == "m" :
            sex_values.append(10)
        elif sex[i] == "f" :
            sex_values.append(-10)
        i+=1
        
    c = np.ones(number_of_users_in_dataset -1) #we make a list c=[1,1,1,...,1] (number_of_users_in_dataset -1 elements which is same size with x , y and q) for constant amount in the equation (ie: cc) 
    x = mean_like_to_follower_for_all_users
    y = ratio_of_followings_to_followers_for_all_users
    q = sex_values


    xx = [] #we want to get xx = x**2
    xxx = [] #we want to get xxx = x**3
    for x_element in x :
        element_power_2 = x_element**2
        xx.append(element_power_2)
        element_power_3 = x_element**3
        xxx.append(element_power_3)
        
    yy = [] #we want to get yy = y**2
    yyy = [] #we want to get yyy = y**3
    for y_element in y :
        element_power_2 = y_element**2
        yy.append(element_power_2)
        element_power_3 = y_element**3
        yyy.append(element_power_3)


        


    
    z = number_of_self_picture_posts_form_9_previous_posts
    A = np.column_stack ((c,x,xx,xxx,y,yy,yyy,q))
    B = z
    result, _ , _ , _ = np.linalg.lstsq(A, B)
    cc , cx , cxx , cxxx, cy , cyy , cyyy , cq   = result
#####################################    print("cc = ",cc , "cx = ",cx , "cxx = ",cxx , "cxxx = ",cxxx , "cy = ",cy , "cyy = ",cyy , "cyyy = ",cyyy , "cq = ",cq )

#----------------------------------------
        # S.H.O.W  R.E.S.A.L.T.S   D.A.T.A.S.E.T   A.F.T.E.R   150   D.A.Y.S "
    
# General Info #
#####################################    print ("number_of_all_users_in_dataset = ", number_of_users_in_dataset - 1)
#####################################    print ("number_of_female_users_in_dataset :",f_number)
#####################################    print ("number_of_male_users_in_dataset :",m_number)
    
# Info based on gender #

#####################################    print ("-------------------------------------------")
#####################################    print ("-------------------------------------------")
#####################################    print ("female stattistics : ")
#####################################    print ("number of female users : " , f_number)
#####################################    print ("mean_of_self_post_for_each_female : " , mean_of_self_post_for_females)
#####################################    print ("mean_number_of_followers_for_each_female : " , mean_number_of_followers_for_females)
#####################################    print ("mean_number_of_followings_for_each_female : " , mean_number_of_followings_for_females)
#####################################    print ("mean_number_of_posts_for_each_female : " , mean_number_of_posts_for_females)
#####################################    print ("-------------------------------------------")
#####################################    print ("-------------------------------------------")
#####################################    print ("male stattistics : ")
#####################################    print ("number of male users : " , m_number)
#####################################    print ("mean_of_self_post_for_each_male : " , mean_of_self_post_for_males)
#####################################    print ("mean_number_of_followers_for_each_male : " , mean_number_of_followers_for_males)
#####################################    print ("mean_number_of_followings_for_each_male : " , mean_number_of_followings_for_males)
#####################################    print ("mean_number_of_posts_for_each_male : " , mean_number_of_posts_for_males)
#####################################    print ("-------------------------------------------")
#####################################    print ("-------------------------------------------")

    return name,number_of_posts,number_of_followers,number_of_followings,mean_like_for_all_users,number_of_self_picture_posts_form_9_previous_posts,sex,narcissism,ratio_of_followings_to_followers_for_all_users,mean_like_to_follower_for_all_users,cc , cx , cxx , cxxx, cy , cyy , cyyy , cq




#-------------------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------


                                  #                    M.A.I.N       D.A.T.A.S.E.T              #
                                  

def test_dataset() :
    import os #for giving adress to where we want to save output files
    from openpyxl import load_workbook
    wb = load_workbook('TestDataset.xlsx')
    ws = wb.active
    name=[]
    number_of_posts = []
    number_of_followers = []
    number_of_followings = []
    number_of_likes_for_10th_ex_post = []
    number_of_likes_for_11th_ex_post = []
    number_of_likes_for_12th_ex_post = []
    number_of_self_picture_posts_form_9_previous_posts = []
    sex = []
    
    #getting data of each column of excel file into a list named by the name of corrospondant name of that column 
    for row in ws :
        if row[0].value in name_150 :  # because we want to compare main dataset with itself after 150 days, and since after 150 days the profile of some users were not availbale so we need to delete those users who are not available in 'main_dataset_after_150_days' by this line 
            name.append(row[0].value)
            number_of_posts.append(row[1].value)
            number_of_followers.append(row[2].value)
            number_of_followings.append(row[3].value)
            number_of_likes_for_10th_ex_post.append(row[4].value)
            number_of_likes_for_11th_ex_post.append(row[5].value)
            number_of_likes_for_12th_ex_post.append(row[6].value)
            number_of_self_picture_posts_form_9_previous_posts.append(row[7].value)
            sex.append(row[8].value)

#------------------------------------
    #finding the number of users in dataset
    number_of_users_in_dataset = len (name)    


#------------------------------------        
    #finding the number of male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    f_number = 0
    m_number = 0
    while i < number_of_users_in_dataset : #because total number of users is 1000 and the first row is for labels so we have to go up to 1001
        if sex[i] == "f" :
            f_number += 1

        elif sex[i] == "m" :
            m_number += 1
        else :
            print("aha")
        i+=1

#------------------------------------
    #finding the mean number of self posts (from 9 posts) for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_self_post_for_all_females = 0
    sum_self_post_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_self_picture_posts_form_9_previous_posts[i]
            sum_self_post_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_self_picture_posts_form_9_previous_posts[i]
            sum_self_post_for_all_males += int(w)
            
        i+=1
        
    mean_of_self_post_for_females = sum_self_post_for_all_females/f_number
    mean_of_self_post_for_males = sum_self_post_for_all_males/m_number

#-----------------------------------
    #finding the mean number of followers for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_followers_for_all_females = 0
    sum_number_of_followers_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_followers[i]
            sum_number_of_followers_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_followers[i]
            sum_number_of_followers_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_followers_for_females = sum_number_of_followers_for_all_females/f_number
    mean_number_of_followers_for_males = sum_number_of_followers_for_all_males/m_number
    
#-------------------------------------
    #finding the mean number of followings for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_followings_for_all_females = 0
    sum_number_of_followings_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_followings[i]
            sum_number_of_followings_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_followings[i]
            sum_number_of_followings_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_followings_for_females = sum_number_of_followings_for_all_females/f_number
    mean_number_of_followings_for_males = sum_number_of_followings_for_all_males/m_number

#-------------------------------------
    #finding the mean number of posts for male and female users in dataset
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    sum_number_of_posts_for_all_females = 0
    sum_number_of_posts_for_all_males = 0
    while i < number_of_users_in_dataset : 
        if sex[i] == "f" :
            t = number_of_posts[i]
            sum_number_of_posts_for_all_females += int(t)
        elif sex[i] == "m" :
            w = number_of_posts[i]
            sum_number_of_posts_for_all_males += int(w)
            
        i+=1
        
    mean_number_of_posts_for_females = sum_number_of_posts_for_all_females/f_number
    mean_number_of_posts_for_males = sum_number_of_posts_for_all_males/m_number
#-------------------------------------
    #finding the relation between number of self posts(in 9 previous posts) and following/follower 
    i=1 #row[0] belongs to lables of columns so by setting i=1 we ignore that first row
    ratio_of_followings_to_followers_for_all_users = []
    while i < number_of_users_in_dataset : 
        t = number_of_followings[i]/number_of_followers[i]
        ratio_of_followings_to_followers_for_all_users.append(t)   
        i+=1
    i = 0

    del number_of_self_picture_posts_form_9_previous_posts[0] #before this command, "number_of_self_picture_posts_form_9_previous_posts" was a list by 1001 elemnts, remember the first element was the label, but we need a that list to just have 1000 elemnts, so be the same size of "ratio_of_followings_to_followers_for_all_users". by having both a 1000 elements list then we can compare and analyse them.

#---------------------------------------------------------
   #finding the relation between the mean number of likes for posts(based on ex 10th 11th 12th posts) and the ratio of following/follower
    i=1
    mean_like_for_all_users = []
    while i < number_of_users_in_dataset :
        total_like = number_of_likes_for_10th_ex_post[i] + number_of_likes_for_11th_ex_post[i] + number_of_likes_for_12th_ex_post[i]
        mean_like = total_like/3
        mean_like_for_all_users.append (mean_like)
        i+=1
##    import matplotlib.pyplot as plt
##    import numpy
##    plt.hist2d(mean_like_for_all_users,ratio_of_followings_to_followers_for_all_users,50)
##    plt.colorbar()
##    plt.show()

#-------- Hist2D : like/following & following/follower
    #finding the relation between the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/following" and the ratio of "following/follower"
    i=0
    mean_like_to_following_for_all_users = []
    del number_of_followings[0]
    while i < number_of_users_in_dataset -1 :
        mean_like_to_following = mean_like_for_all_users[i]/number_of_followings[i]
        mean_like_to_following_for_all_users.append (mean_like_to_following)
        i+=1
    import matplotlib.pyplot as plt
    import numpy
#####################################    print("MAX[mean_like_to_following_for_all_users] = " , max(mean_like_to_following_for_all_users)) #we need this line to set max limit for horizental axis in plt.hist2d command (two lines below), which is 32.12 (we set to 33 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
#####################################    print("MAX[ratio_of_followings_to_followers_for_all_users] = " , max(ratio_of_followings_to_followers_for_all_users)) #we need this line to set max limit for vertical axis in plt.hist2d command (two lines below), which is 15.66 (we set to 16 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
    plt.hist2d(mean_like_to_following_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,2],[0,16]]) #notice max horizental was 33 but since there was only a few data with more than 2 as ratio of like to following so we ignored them and set the max horizental limit to 2
##    plt.scatter(mean_like_to_following_for_all_users,ratio_of_followings_to_followers_for_all_users)
    plt.xlabel('Like/Following ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Following & Following/Follower',fontsize=12)
    file_name = "Hist2D_LikeFollowing_to_FollowingFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
#-------- Hist2D : like/follower & following/follower
    #finding the relation between the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/follower" and the ratio of "following/follower"
    i=0
    mean_like_to_follower_for_all_users = []
    del number_of_followers[0]
    while i < number_of_users_in_dataset - 1 :
        mean_like_to_follower = mean_like_for_all_users[i]/number_of_followers[i]
        mean_like_to_follower_for_all_users.append (mean_like_to_follower)
        i+=1
    import matplotlib.pyplot as plt
    import numpy as np
    import matplotlib.colors as mcolors
    import matplotlib.mlab as mlab
    import os #To give address for saving output plots
##    import scipy.stats

##    from numpy._distributor_init import NUMPY_MKL 
##    from scipy.stats import norm
#####################################    print("MAX[mean_like_to_follower_for_all_users] = " , max(mean_like_to_follower_for_all_users)) #we need this line to set max limit for horizental axis in plt.hist2d command (two lines below), which is 0.92 (we set to 1 ) , notice once we run the program and see the 0.92 then we set it in plt.hist2d for ever
#####################################    print("MAX[ratio_of_followings_to_followers_for_all_users] = " , max(ratio_of_followings_to_followers_for_all_users)) #we need this line to set max limit for vertical axis in plt.hist2d command (two lines below), which is 15.66 (we set to 16 ) , notice once we run the program and see the 15.66 then we set it in plt.hist2d for ever
    plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,1],[0,16]])
##    counts, _ , _ , _ = plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,100,[[0,1],[0,16]])
##    row_total = 0
##    total = 0
##    for w in counts :
##        row_total += w
##    for y in row_total :
##        total += y
##    print(total)
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Follower & Following/Follower',fontsize=12)
    file_name = "Hist2D_LikeFollower_to_FollowingFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #---- Hist2D : like/follower & following/follower --- its the above hist2d but here we just focus on the dense area  
    plt.hist2d(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users,30,[[0,0.5],[0,10]])
##    plt.scatter(mean_like_to_follower_for_all_users,ratio_of_followings_to_followers_for_all_users)
##    plt.show()
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Following/Follower ratio')
##    plt.title('Like/Follower & Following/Follower (focused)',fontsize=12)
    file_name = "Hist2Dfocused_LikeFollower_to_FollowingFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !

    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 


    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #-------- Hist1D : following/follower 
    plt.hist(ratio_of_followings_to_followers_for_all_users,[0,0.5,1,1.5,2,2.5,3,3.5,4,4.5,5,5.5,6,6.5,7,7.5,8,8.5,9,9.5,10,10.5,11,11.5,12,12.5,13,13.5,14,14.5,15,15.5,16]) #we set the edges of bins for histogram from 0 to 16(which is the max in the dataset and we noticed that from previous command  

    #-----
##    pdf_x = np.linspace(np.min(ratio_of_followings_to_followers_for_all_users),np.max(ratio_of_followings_to_followers_for_all_users),100)
##    pdf_y = 1.0/np.sqrt(2*np.pi*var)*np.exp(-0.5*(pdf_x-avg)**2/var)
##    plt.plot(pdf_x,pdf_y,'k--')
##    print(avg)
##    print(var)
##    plt.colorbar()
##    (mu, sigma) = norm.fit(datos)
##    y = mlab.normpdf( bins, mu, sigma)
##    l = plt.plot(bins, y, 'r--', linewidth=2)
     
    plt.xlabel('Following/Follower ratio')
    plt.ylabel('Frequency')
##    plt.title('Following/Follower',fontsize=12)
    file_name = "Hist1D_Following_to_Follower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #-------- Hist1D : like/follower

    plt.hist(mean_like_to_follower_for_all_users,[0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]) #we set the edges of bins for histogram from 0 to 0.92(which is the max in the dataset and we noticed that from previous commands  
    #--------
##    pdf_x = np.linspace(np.min(ratio_of_followings_to_followers_for_all_users),np.max(ratio_of_followings_to_followers_for_all_users),100)
##    pdf_y = 1.0/np.sqrt(2*np.pi*var)*np.exp(-0.5*(pdf_x-avg)**2/var)
##    plt.plot(pdf_x,pdf_y,'k--')
##    print(avg)
##    print(var)
##    plt.colorbar()
##    (mu, sigma) = norm.fit(datos)
##    y = mlab.normpdf( bins, mu, sigma)
##    l = plt.plot(bins, y, 'r--', linewidth=2)
     
    plt.xlabel('Like/Follower ratio')
    plt.ylabel('Frequency')
##    plt.title('Like/Follower',fontsize=12)
    file_name = "Hist1D_Like_to_Follower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #-------- 

    

#-----------------------------------------------------------
    #great result :
    #THOSE USERS WITH FOLLOWEING/FOLLOWER < 1 WITH STRONGER PROBABILITY HAVE BETTER CHANCE OF ACCEPTANCE . In other words they have better like/follower (like/following) , so the chance that they start the rumor are more stronger 
#-----------------------------------------------------------    
    #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the ratio of "mean number of likes for posts(based on ex 10th 11th 12th posts)/follower"
    #PLEASE notice that if we use discrete values of 0,1,2,3,..,9 as number of self posts(narcissism) then hist2d shape would be discrete that is not good, so for each number we randomly choose a value between two defined limit. for example if self posts number is 2 then we choose ranmoly a number between (0.2 , 0.3) to represent its narcissism level
    i=0
    import matplotlib.pyplot as plt
    import numpy
    import random
    narcissism = []
    while i < number_of_users_in_dataset - 1 :
        if number_of_self_picture_posts_form_9_previous_posts[i] == 0 :
            t = random.uniform(0,0.1)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 1 :
            t = random.uniform(0.1,0.2)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 2 :
            t = random.uniform(0.2,0.3)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 3 :
            t = random.uniform(0.3,0.4)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 4 :
            t = random.uniform(0.4,0.5)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 5 :
            t = random.uniform(0.5,0.6)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 6 :
            t = random.uniform(0.6,0.7)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 7 :
            t = random.uniform(0.7,0.8)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 8 :
            t = random.uniform(0.8,0.9)
            narcissism.append (t)
        elif number_of_self_picture_posts_form_9_previous_posts[i] == 9 :
            t = random.uniform(0.9,1)
            narcissism.append (t)
        i+=1
    plt.hist2d(narcissism,mean_like_to_follower_for_all_users,8,[[0,1],[0,1]])
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('Like/Follower')
##    plt.title('narcissism & Like/Follower ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_LikeFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !

    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
#---- #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the ratio of following/follower"
    plt.hist2d(narcissism,ratio_of_followings_to_followers_for_all_users,8,[[0,1],[0,16]])
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('Following/Follower')
##    plt.title('narcissism & Following/Follower ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_FollowingFollower"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
##----------
    #---- #finding the relation between the number of self posts(in 9 previous posts) (NARCISSISM) and the number of posts
    i=1
    del number_of_posts[0]
    plt.hist2d(narcissism,number_of_posts,40,[[0,1],[50,2000]]) #notice the min of posts was 50 because we didnt pick users with less than 50 posts and also max was 10000 but since there was a few users with more than 2000 so we made up limit to 2000
    plt.xlabel('Ratio of self-presenting posts')
    plt.ylabel('number_of_posts')
##    plt.title('narcissism & number_of_posts ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_numberofposts"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()
    #-----
    #---- #finding the following/follower ratio and the number of posts
    i=1
    plt.hist2d(ratio_of_followings_to_followers_for_all_users,number_of_posts,40,[[0,16],[50,2000]]) #notice the min of posts was 50 because we didnt pick users with less than 50 posts and also max was 10000 but since there was a few users with more than 2000 so we made up limit to 2000
    plt.xlabel('following/follower')
    plt.ylabel('number_of_posts')
    plt.title('following/follower & number_of_posts ratio',fontsize=12)
    file_name = "Hist2D_narcissism_to_numberofposts"  # notice that if you set name as following/follower , I mean if you put / symbol in the name then it won't be saved! because symboles such this "/" must not be in file names
    path_of_saving = str ("C:\\Users\\novin\\Desktop\\Instagram_Analysis\\Compare two main dataset (the main dataset and the main dataset after 150 days)\\Result_MainDataset_Pics") #here we make the address of the folder we want pictures to be saved, BUT YOU HAVE TO douplicate each Backslash in the URL address (as you see I did ) , if you dont it wont work and give you error of UNICODESCAPE !
    cbar = plt.colorbar() #the reason why we don't use only plt.colorbar() lonely , is that we want to put a label close it, so we used this line and next line 
    cbar.set_label ("Frequency")
    plt.savefig(os.path.join(path_of_saving, file_name))
    plt.close() ##########################################################################################
#####################################    plt.show()

    

# finding the relation of self post with  "following/follower" , "like/follower" and "sex"
# notic we want to find best amount for cc , cx , cxx , cy , cyy , cq in the equation z = cc + cx*x + cxx*x^2 + cxxx*x^3 + cy*y + cyy*y^2 + cyyy*y^3 + cq*q (in which x , y , q are independent variables)
    import matplotlib.pyplot as plt
    import numpy as np
    i = 1
    sex_values = []
    while i < number_of_users_in_dataset :
        if sex[i] == "m" :
            sex_values.append(10)
        elif sex[i] == "f" :
            sex_values.append(-10)
        i+=1
        
    c = np.ones(number_of_users_in_dataset -1) #we make a list c=[1,1,1,...,1] (number_of_users_in_dataset -1 elements which is same size with x , y and q) for constant amount in the equation (ie: cc) 
    x = mean_like_to_follower_for_all_users
    y = ratio_of_followings_to_followers_for_all_users
    q = sex_values


    xx = [] #we want to get xx = x**2
    xxx = [] #we want to get xxx = x**3
    for x_element in x :
        element_power_2 = x_element**2
        xx.append(element_power_2)
        element_power_3 = x_element**3
        xxx.append(element_power_3)
        
    yy = [] #we want to get yy = y**2
    yyy = [] #we want to get yyy = y**3
    for y_element in y :
        element_power_2 = y_element**2
        yy.append(element_power_2)
        element_power_3 = y_element**3
        yyy.append(element_power_3)


        


    
    z = number_of_self_picture_posts_form_9_previous_posts
    A = np.column_stack ((c,x,xx,xxx,y,yy,yyy,q))
    B = z
    result, _ , _ , _ = np.linalg.lstsq(A, B)
    cc , cx , cxx , cxxx, cy , cyy , cyyy , cq   = result
#####################################    print("cc = ",cc , "cx = ",cx , "cxx = ",cxx , "cxxx = ",cxxx , "cy = ",cy , "cyy = ",cyy , "cyyy = ",cyyy , "cq = ",cq )

#----------------------------------------
        # S.H.O.W  R.E.S.A.L.T.S   D.A.T.A.S.E.T "
    
# General Info #
#####################################    print("\n\n\n\n\n\n")
#####################################    print ("number_of_all_users_in_dataset = ", number_of_users_in_dataset - 1)
#####################################    print ("number_of_female_users_in_dataset :",f_number)
#####################################    print ("number_of_male_users_in_dataset :",m_number)
    
# Info based on gender #

#####################################    print ("-------------------------------------------")
#####################################    print ("-------------------------------------------")
#####################################    print ("female stattistics : ")
#####################################    print ("number of female users : " , f_number)
#####################################    print ("mean_of_self_post_for_each_female : " , mean_of_self_post_for_females)
#####################################    print ("mean_number_of_followers_for_each_female : " , mean_number_of_followers_for_females)
#####################################    print ("mean_number_of_followings_for_each_female : " , mean_number_of_followings_for_females)
#####################################    print ("mean_number_of_posts_for_each_female : " , mean_number_of_posts_for_females)
#####################################    print ("-------------------------------------------")
#####################################    print ("-------------------------------------------")
#####################################    print ("male stattistics : ")
#####################################    print ("number of male users : " , m_number)
#####################################    print ("mean_of_self_post_for_each_male : " , mean_of_self_post_for_males)
#####################################    print ("mean_number_of_followers_for_each_male : " , mean_number_of_followers_for_males)
#####################################    print ("mean_number_of_followings_for_each_male : " , mean_number_of_followings_for_males)
#####################################    print ("mean_number_of_posts_for_each_male : " , mean_number_of_posts_for_males)
#####################################    print ("-------------------------------------------")
#####################################    print ("-------------------------------------------")

    return name,number_of_posts,number_of_followers,number_of_followings,mean_like_for_all_users,number_of_self_picture_posts_form_9_previous_posts,sex,narcissism,ratio_of_followings_to_followers_for_all_users,mean_like_to_follower_for_all_users,cc , cx , cxx , cxxx, cy , cyy , cyyy , cq


#-------------------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------

def followers_growth (number_of_followers_150,number_of_followers) :
    fr_growth = []
    number_of_users = len (number_of_followers_150)
    i = 0
    while i < number_of_users :
        j = number_of_followers_150[i] - number_of_followers [i]
        fr_growth.append(j)
        i += 1
#    print (fr_growth)
    return fr_growth
        
#--------------------------------------------------------------------------------
def followings_growth (number_of_followings_150,number_of_followings) :
    fw_growth = []
    number_of_users = len (number_of_followings_150)
    i = 0
    while i < number_of_users :
        j = number_of_followings_150[i] - number_of_followings [i]
        fw_growth.append(j)
        i += 1
#    print (fw_growth)
    return fw_growth
#--------------------------------------------------------------------------------
def like_growth (mean_like_for_all_users_150,mean_like_for_all_users) :
    li_growth = []
    number_of_users = len (mean_like_for_all_users_150)
    i = 0
    while i < number_of_users :
        j = mean_like_for_all_users_150[i] - mean_like_for_all_users[i]
        li_growth.append(j)
        i += 1
#    print (li_growth)
    return li_growth
#--------------------------------------------------------------------------------
def self_presenting_growth (number_of_self_picture_posts_form_9_previous_posts_150,number_of_self_picture_posts_form_9_previous_posts) :
    sp_growth = []
    number_of_users = len (mean_like_for_all_users_150)
    i = 0
    while i < number_of_users :
        j = number_of_self_picture_posts_form_9_previous_posts_150[i] - number_of_self_picture_posts_form_9_previous_posts[i]
        sp_growth.append(j)
        i += 1
#    print (sp_growth)
    return sp_growth

#--------------------------------------------------------------------------------
def number_of_posts_growth (number_of_posts_150,number_of_posts) :
    np_growth = []
    number_of_users = len (number_of_posts_150)
    i = 0
    while i < number_of_users :
        j = number_of_posts_150[i] - number_of_posts[i]
        np_growth.append(j)
        i += 1

    return np_growth
#--------------------------------------------------------------------------------
def dynamic_growth (narcissism_150,ratio_of_followings_to_followers_for_all_users_150, mean_like_to_follower_for_all_users_150,narcissism,ratio_of_followings_to_followers_for_all_users,mean_like_to_follower_for_all_users):
    number_of_users = len (narcissism_150)
    n_growth = []
    fwtofr_growth = []
    litofr_growth = []
    i = 0
    while i < number_of_users :
        j = narcissism_150[i] - narcissism[i]
        n_growth.append(j)
        k = ratio_of_followings_to_followers_for_all_users_150[i] - ratio_of_followings_to_followers_for_all_users[i]
        fwtofr_growth.append(k)
        z = mean_like_to_follower_for_all_users_150[i] - mean_like_to_follower_for_all_users[i]
        litofr_growth.append(z)
        i += 1

    return n_growth,fwtofr_growth,litofr_growth
#----------------------------------------------------------------------------------        


                                #          Compare      Main_dataset_after_150_days       with         Main_dataset             #

name_150,number_of_posts_150,number_of_followers_150,number_of_followings_150,mean_like_for_all_users_150,number_of_self_picture_posts_form_9_previous_posts_150,sex_150,narcissism_150,ratio_of_followings_to_followers_for_all_users_150, mean_like_to_follower_for_all_users_150 ,cc_150 , cx_150 , cxx_150 , cxxx_150, cy_150 , cyy_150 , cyyy_150 , cq_150 = test_after_150_days()
name,number_of_posts,number_of_followers,number_of_followings,mean_like_for_all_users,number_of_self_picture_posts_form_9_previous_posts,sex,narcissism,ratio_of_followings_to_followers_for_all_users,mean_like_to_follower_for_all_users,cc , cx , cxx , cxxx, cy , cyy , cyyy , cq = test_dataset()
fr = followers_growth (number_of_followers_150,number_of_followers)
fw = followings_growth (number_of_followings_150,number_of_followings)
lg = like_growth (mean_like_for_all_users_150,mean_like_for_all_users)
spg = self_presenting_growth (number_of_self_picture_posts_form_9_previous_posts_150,number_of_self_picture_posts_form_9_previous_posts)
np_growth = number_of_posts_growth (number_of_posts_150,number_of_posts)
n_growth,fwtofr_growth,litofr_growth = dynamic_growth (narcissism_150,ratio_of_followings_to_followers_for_all_users_150, mean_like_to_follower_for_all_users_150,narcissism,ratio_of_followings_to_followers_for_all_users,mean_like_to_follower_for_all_users)

import numpy
from scipy.stats import pearsonr
from scipy.stats import spearmanr
from scipy.stats import kendalltau

# pearson correlation 
fwtofr_growth_to_fwtofr = pearsonr(fwtofr_growth,ratio_of_followings_to_followers_for_all_users)
fwtofr_growth_to_litofr = pearsonr(fwtofr_growth,mean_like_to_follower_for_all_users)
fwtofr_growth_to_narcissism = pearsonr(fwtofr_growth,narcissism)
fwtofr_growth_to_number_of_posts = pearsonr(fwtofr_growth,number_of_posts)

litofr_growth_to_fwtofr = pearsonr(litofr_growth,ratio_of_followings_to_followers_for_all_users)
litofr_growth_to_litofr = pearsonr(litofr_growth,mean_like_to_follower_for_all_users)
litofr_growth_to_narcissism = pearsonr(litofr_growth,narcissism)
litofr_growth_to_number_of_posts = pearsonr(litofr_growth,number_of_posts)

n_growth_to_fwtofr = pearsonr(n_growth,ratio_of_followings_to_followers_for_all_users)
n_growth_to_litofr = pearsonr(n_growth,mean_like_to_follower_for_all_users)
n_growth_to_narcissism = pearsonr(n_growth,narcissism)
n_growth_to_number_of_posts = pearsonr(n_growth,number_of_posts)

np_growth_to_fwtofr = pearsonr(np_growth,ratio_of_followings_to_followers_for_all_users)
np_growth_to_litofr = pearsonr(np_growth,mean_like_to_follower_for_all_users)
np_growth_to_narcissism = pearsonr(np_growth,narcissism)
np_growth_to_number_of_posts = pearsonr(np_growth,number_of_posts)

# spearman correlation 
fwtofr_growth_to_fwtofr_s = spearmanr(fwtofr_growth,ratio_of_followings_to_followers_for_all_users)
fwtofr_growth_to_litofr_s = spearmanr(fwtofr_growth,mean_like_to_follower_for_all_users)
fwtofr_growth_to_narcissism_s = spearmanr(fwtofr_growth,narcissism)
fwtofr_growth_to_number_of_posts_s = spearmanr(fwtofr_growth,number_of_posts)

litofr_growth_to_fwtofr_s = spearmanr(litofr_growth,ratio_of_followings_to_followers_for_all_users)
litofr_growth_to_litofr_s = spearmanr(litofr_growth,mean_like_to_follower_for_all_users)
litofr_growth_to_narcissism_s = spearmanr(litofr_growth,narcissism)
litofr_growth_to_number_of_posts_s = spearmanr(litofr_growth,number_of_posts)

n_growth_to_fwtofr_s = spearmanr(n_growth,ratio_of_followings_to_followers_for_all_users)
n_growth_to_litofr_s = spearmanr(n_growth,mean_like_to_follower_for_all_users)
n_growth_to_narcissism_s = spearmanr(n_growth,narcissism)
n_growth_to_number_of_posts_s = spearmanr(n_growth,number_of_posts)

np_growth_to_fwtofr_s = spearmanr(np_growth,ratio_of_followings_to_followers_for_all_users)
np_growth_to_litofr_s = spearmanr(np_growth,mean_like_to_follower_for_all_users)
np_growth_to_narcissism_s = spearmanr(np_growth,narcissism)
np_growth_to_number_of_posts_s = spearmanr(np_growth,number_of_posts)

# kendall correlation 
fwtofr_growth_to_fwtofr_k = kendalltau(fwtofr_growth,ratio_of_followings_to_followers_for_all_users)
fwtofr_growth_to_litofr_k = kendalltau(fwtofr_growth,mean_like_to_follower_for_all_users)
fwtofr_growth_to_narcissism_k = kendalltau(fwtofr_growth,narcissism)
fwtofr_growth_to_number_of_posts_k = kendalltau(fwtofr_growth,number_of_posts)

litofr_growth_to_fwtofr_k = kendalltau(litofr_growth,ratio_of_followings_to_followers_for_all_users)
litofr_growth_to_litofr_k = kendalltau(litofr_growth,mean_like_to_follower_for_all_users)
litofr_growth_to_narcissism_k = kendalltau(litofr_growth,narcissism)
litofr_growth_to_number_of_posts_k = kendalltau(litofr_growth,number_of_posts)

n_growth_to_fwtofr_k = kendalltau(n_growth,ratio_of_followings_to_followers_for_all_users)
n_growth_to_litofr_k = kendalltau(n_growth,mean_like_to_follower_for_all_users)
n_growth_to_narcissism_k = kendalltau(n_growth,narcissism)
n_growth_to_number_of_posts_k = kendalltau(n_growth,number_of_posts)

np_growth_to_fwtofr_k = kendalltau(np_growth,ratio_of_followings_to_followers_for_all_users)
np_growth_to_litofr_k = kendalltau(np_growth,mean_like_to_follower_for_all_users)
np_growth_to_narcissism_k = kendalltau(np_growth,narcissism)
np_growth_to_number_of_posts_k = kendalltau(np_growth,number_of_posts)


########
########
########print(" fwtofr_growth_to_fwtofrs : ", fwtofr_growth_to_fwtofr)
########print(" fwtofr_growth_to_litofr : ", fwtofr_growth_to_litofr)
########print(" fwtofr_growth_to_narcissism : ", fwtofr_growth_to_narcissism)
########
########print(" litofr_growth_to_fwtofr : ", litofr_growth_to_fwtofr)
########print(" litofr_growth_to_litofr : ", litofr_growth_to_litofr)
########print(" litofr_growth_to_narcissism : ", litofr_growth_to_narcissism)
########
########print(" n_growth_to_fwtofr : ", n_growth_to_fwtofr)
########print(" n_growth_to_litofr : ", n_growth_to_litofr)
########print(" n_growth_to_narcissism : ", n_growth_to_narcissism)

n_growth_to_fwtofr_growth = pearsonr(n_growth,fwtofr_growth)
n_growth_to_litofr_growth = pearsonr(n_growth,litofr_growth)
fwtofr_growth_to_litofr_growth = pearsonr(fwtofr_growth,litofr_growth)
np_growth_to_n_growth = pearsonr(np_growth,n_growth)
np_growth_to_litofr_growth = pearsonr(np_growth,litofr_growth)
np_growth_to_fwtofr_growth = pearsonr(np_growth,fwtofr_growth)

n_growth_to_fwtofr_growth_s = spearmanr(n_growth,fwtofr_growth)
n_growth_to_litofr_growth_s = spearmanr(n_growth,litofr_growth)
fwtofr_growth_to_litofr_growth_s = spearmanr(fwtofr_growth,litofr_growth)
np_growth_to_n_growth_s = spearmanr(np_growth,n_growth)
np_growth_to_litofr_growth_s = spearmanr(np_growth,litofr_growth)
np_growth_to_fwtofr_growth_s = spearmanr(np_growth,fwtofr_growth)

n_growth_to_fwtofr_growth_k = kendalltau(n_growth,fwtofr_growth)
n_growth_to_litofr_growth_k = kendalltau(n_growth,litofr_growth)
fwtofr_growth_to_litofr_growth_k = kendalltau(fwtofr_growth,litofr_growth)
np_growth_to_n_growth_k = kendalltau(np_growth,n_growth)
np_growth_to_litofr_growth_k = kendalltau(np_growth,litofr_growth)
np_growth_to_fwtofr_growth_k = kendalltau(np_growth,fwtofr_growth)

#------------------------
#print(fwtofr_growth)
import statistics

mean_fwtofr = statistics.mean(ratio_of_followings_to_followers_for_all_users)
stdv_fwtofr = statistics.stdev(ratio_of_followings_to_followers_for_all_users)

mean_fwtofr_150 = statistics.mean(ratio_of_followings_to_followers_for_all_users_150)
stdv_fwtofr_150 = statistics.stdev(ratio_of_followings_to_followers_for_all_users_150)

mean_litofr = statistics.mean(mean_like_to_follower_for_all_users)
stdv_litofr = statistics.stdev(mean_like_to_follower_for_all_users)

mean_litofr_150 = statistics.mean(mean_like_to_follower_for_all_users_150)
stdv_litofr_150 = statistics.stdev(mean_like_to_follower_for_all_users_150)

mean_nar = statistics.mean(narcissism)
stdv_nar = statistics.stdev(narcissism)

mean_nar_150 = statistics.mean(narcissism_150)
stdv_nar_150 = statistics.stdev(narcissism_150)

mean_np = statistics.mean(number_of_posts)
stdv_np = statistics.stdev(number_of_posts)

mean_np_150 = statistics.mean(number_of_posts_150)
stdv_np_150 = statistics.stdev(number_of_posts_150)
#----------------
import heapq
iii = heapq.nlargest(15, range(len(np_growth)), np_growth.__getitem__)  

fwtofr_selected_largest_np_growth = []
litofr_selected_largest_np_growth = []
narci_selected_largest_np_growth = []
number_of_posts_selected_largest_np_growth = []
name_largest_np_growth = []

for i in iii:
    fwtofr_selected_largest_np_growth.append(ratio_of_followings_to_followers_for_all_users[i])
    litofr_selected_largest_np_growth.append(mean_like_to_follower_for_all_users[i])
    narci_selected_largest_np_growth.append(narcissism[i])
    number_of_posts_selected_largest_np_growth.append(number_of_posts[i])
    name_largest_np_growth.append(name[i+1])#row 0 in names list belongs to labels so we add 1 to i

mean_fwtofr_selected_largest_np_growth = statistics.mean(fwtofr_selected_largest_np_growth)
stdv_fwtofr_selected_largest_np_growth = statistics.stdev(fwtofr_selected_largest_np_growth)
mean_litofr_selected_largest_np_growth = statistics.mean(litofr_selected_largest_np_growth)
stdv_litofr_selected_largest_np_growth = statistics.stdev(litofr_selected_largest_np_growth)
mean_narci_selected_largest_np_growth = statistics.mean(narci_selected_largest_np_growth)
stdv_narci_selected_largest_np_growth = statistics.stdev(narci_selected_largest_np_growth)
mean_number_of_posts_selected_largest_np_growth = statistics.mean(number_of_posts_selected_largest_np_growth)
stdv_number_of_posts_selected_largest_np_growth = statistics.stdev(number_of_posts_selected_largest_np_growth)
#-----------------
eee = heapq.nsmallest(15, range(len(np_growth)), np_growth.__getitem__)  

fwtofr_selected_smallest_np_growth = []
litofr_selected_smallest_np_growth = []
narci_selected_smallest_np_growth = []
number_of_posts_selected_smallest_np_growth = []
name_smallest_np_growth = []

for i in eee:
    fwtofr_selected_smallest_np_growth.append(ratio_of_followings_to_followers_for_all_users[i])
    litofr_selected_smallest_np_growth.append(mean_like_to_follower_for_all_users[i])
    narci_selected_smallest_np_growth.append(narcissism[i])
    number_of_posts_selected_smallest_np_growth.append(number_of_posts[i])
    name_smallest_np_growth.append(name[i+1]) #row 0 in names list belongs to labels so we add 1 to i

mean_fwtofr_selected_smallest_np_growth = statistics.mean(fwtofr_selected_smallest_np_growth)
stdv_fwtofr_selected_smallest_np_growth = statistics.stdev(fwtofr_selected_smallest_np_growth)
mean_litofr_selected_smallest_np_growth = statistics.mean(litofr_selected_smallest_np_growth)
stdv_litofr_selected_smallest_np_growth = statistics.stdev(litofr_selected_smallest_np_growth)
mean_narci_selected_smallest_np_growth = statistics.mean(narci_selected_smallest_np_growth)
stdv_narci_selected_smallest_np_growth = statistics.stdev(narci_selected_smallest_np_growth)
mean_number_of_posts_selected_smallest_np_growth = statistics.mean(number_of_posts_selected_smallest_np_growth)
stdv_number_of_posts_selected_smallest_np_growth = statistics.stdev(number_of_posts_selected_smallest_np_growth)
#-----------------
jjj = heapq.nlargest(15, range(len(litofr_growth)), litofr_growth.__getitem__)  

fwtofr_selected_largest_litofr_growth = []
litofr_selected_largest_litofr_growth = []
narci_selected_largest_litofr_growth = []
number_of_posts_selected_largest_litofr_growth = []
name_largest_litofr_growth = []

for i in jjj:
    fwtofr_selected_largest_litofr_growth.append(ratio_of_followings_to_followers_for_all_users[i])
    litofr_selected_largest_litofr_growth.append(mean_like_to_follower_for_all_users[i])
    narci_selected_largest_litofr_growth.append(narcissism[i])
    number_of_posts_selected_largest_litofr_growth.append(number_of_posts[i])
    name_largest_litofr_growth.append(name[i+1]) #row 0 in names list belongs to labels so we add 1 to i
    

mean_fwtofr_selected_largest_litofr_growth = statistics.mean(fwtofr_selected_largest_litofr_growth)
stdv_fwtofr_selected_largest_litofr_growth = statistics.stdev(fwtofr_selected_largest_litofr_growth)
mean_litofr_selected_largest_litofr_growth = statistics.mean(litofr_selected_largest_litofr_growth)
stdv_litofr_selected_largest_litofr_growth = statistics.stdev(litofr_selected_largest_litofr_growth)
mean_narci_selected_largest_litofr_growth = statistics.mean(narci_selected_largest_litofr_growth)
stdv_narci_selected_largest_litofr_growth = statistics.stdev(narci_selected_largest_litofr_growth)
mean_number_of_posts_selected_largest_litofr_growth = statistics.mean(number_of_posts_selected_largest_litofr_growth)
stdv_number_of_posts_selected_largest_litofr_growth = statistics.stdev(number_of_posts_selected_largest_litofr_growth)
#-----------------
uuu = heapq.nsmallest(15, range(len(litofr_growth)), litofr_growth.__getitem__)  

fwtofr_selected_smallest_litofr_growth = []
litofr_selected_smallest_litofr_growth = []
narci_selected_smallest_litofr_growth = []
number_of_posts_selected_smallest_litofr_growth = []
name_smallest_litofr_growth = []

for i in uuu:
    fwtofr_selected_smallest_litofr_growth.append(ratio_of_followings_to_followers_for_all_users[i])
    litofr_selected_smallest_litofr_growth.append(mean_like_to_follower_for_all_users[i])
    narci_selected_smallest_litofr_growth.append(narcissism[i])
    number_of_posts_selected_smallest_litofr_growth.append(number_of_posts[i])
    name_smallest_litofr_growth.append(name[i+1])#row 0 in names list belongs to labels so we add 1 to i

mean_fwtofr_selected_smallest_litofr_growth = statistics.mean(fwtofr_selected_smallest_litofr_growth)
stdv_fwtofr_selected_smallest_litofr_growth = statistics.stdev(fwtofr_selected_smallest_litofr_growth)
mean_litofr_selected_smallest_litofr_growth = statistics.mean(litofr_selected_smallest_litofr_growth)
stdv_litofr_selected_smallest_litofr_growth = statistics.stdev(litofr_selected_smallest_litofr_growth)
mean_narci_selected_smallest_litofr_growth = statistics.mean(narci_selected_smallest_litofr_growth)
stdv_narci_selected_smallest_litofr_growth = statistics.stdev(narci_selected_smallest_litofr_growth)
mean_number_of_posts_selected_smallest_litofr_growth = statistics.mean(number_of_posts_selected_smallest_litofr_growth)
stdv_number_of_posts_selected_smallest_litofr_growth = statistics.stdev(number_of_posts_selected_smallest_litofr_growth)
#-----------------
ddd = heapq.nlargest(15, range(len(fwtofr_growth)), fwtofr_growth.__getitem__)  

fwtofr_selected_largest_fwtofr_growth = []
litofr_selected_largest_fwtofr_growth = []
narci_selected_largest_fwtofr_growth = []
number_of_posts_selected_largest_fwtofr_growth = []
name_largest_fwtofr_growth = []
fwtofr_growth_selected = []

for i in ddd:
    fwtofr_selected_largest_fwtofr_growth.append(ratio_of_followings_to_followers_for_all_users[i])
    litofr_selected_largest_fwtofr_growth.append(mean_like_to_follower_for_all_users[i])
    narci_selected_largest_fwtofr_growth.append(narcissism[i])
    number_of_posts_selected_largest_fwtofr_growth.append(number_of_posts[i])
    name_largest_fwtofr_growth.append(name[i+1]) #row 0 in names list belongs to labels so we add 1 to i
    fwtofr_growth_selected.append(fwtofr_growth[i])
##    print ("ratio_of_followings_to_followers_for_all_users[i] : ",ratio_of_followings_to_followers_for_all_users[i])
##    print ("ratio_of_followings_to_followers_for_all_users[i] : ",ratio_of_followings_to_followers_for_all_users_150[i])
##    print ("fwtofr_growth [i] : ",fwtofr_growth [i])
##    print("-----------------------")
##fwtofr_growth_to_fwtofr_selected = numpy.corrcoef(fwtofr_growth_selected,fwtofr_selected_largest_fwtofr_growth)
##print ( "fwtofr_growth_to_fwtofr_selected : ",fwtofr_growth_to_fwtofr_selected)

mean_fwtofr_selected_largest_fwtofr_growth = statistics.mean(fwtofr_selected_largest_fwtofr_growth)
stdv_fwtofr_selected_largest_fwtofr_growth = statistics.stdev(fwtofr_selected_largest_fwtofr_growth)
mean_litofr_selected_largest_fwtofr_growth = statistics.mean(litofr_selected_largest_fwtofr_growth)
stdv_litofr_selected_largest_fwtofr_growth = statistics.stdev(litofr_selected_largest_fwtofr_growth)
mean_narci_selected_largest_fwtofr_growth = statistics.mean(narci_selected_largest_fwtofr_growth)
stdv_narci_selected_largest_fwtofr_growth = statistics.stdev(narci_selected_largest_fwtofr_growth)
mean_number_of_posts_selected_largest_fwtofr_growth = statistics.mean(number_of_posts_selected_largest_fwtofr_growth)
stdv_number_of_posts_selected_largest_fwtofr_growth = statistics.stdev(number_of_posts_selected_largest_fwtofr_growth)
#-----------------
www = heapq.nsmallest(15, range(len(fwtofr_growth)), fwtofr_growth.__getitem__)  

fwtofr_selected_smallest_fwtofr_growth = []
litofr_selected_smallest_fwtofr_growth = []
narci_selected_smallest_fwtofr_growth = []
number_of_posts_selected_smallest_fwtofr_growth = []
name_smallest_fwtofr_growth = []
fwtofr_growth_s_selected = []

for i in www:
    fwtofr_selected_smallest_fwtofr_growth.append(ratio_of_followings_to_followers_for_all_users[i])
    litofr_selected_smallest_fwtofr_growth.append(mean_like_to_follower_for_all_users[i])
    narci_selected_smallest_fwtofr_growth.append(narcissism[i])
    number_of_posts_selected_smallest_fwtofr_growth.append(number_of_posts[i])
    name_smallest_fwtofr_growth.append(name[i+1]) #row 0 in names list belongs to labels so we add 1 to i
    fwtofr_growth_s_selected.append(fwtofr_growth[i])
##    print ("ratio_of_followings_to_followers_for_all_users[i] : ",ratio_of_followings_to_followers_for_all_users[i])
##    print ("ratio_of_followings_to_followers_for_all_users[i] : ",ratio_of_followings_to_followers_for_all_users_150[i])
##    print ("fwtofr_growth [i] : ",fwtofr_growth [i])
##    print("-----------------------")
##fwtofr_growth_to_fwtofr_s_selected = numpy.corrcoef(fwtofr_growth_s_selected,fwtofr_selected_smallest_fwtofr_growth)
##print ( "fwtofr_growth_to_fwtofr_s_selected : ",fwtofr_growth_to_fwtofr_s_selected)

mean_fwtofr_selected_smallest_fwtofr_growth = statistics.mean(fwtofr_selected_smallest_fwtofr_growth)
stdv_fwtofr_selected_smallest_fwtofr_growth = statistics.stdev(fwtofr_selected_smallest_fwtofr_growth)
mean_litofr_selected_smallest_fwtofr_growth = statistics.mean(litofr_selected_smallest_fwtofr_growth)
stdv_litofr_selected_smallest_fwtofr_growth = statistics.stdev(litofr_selected_smallest_fwtofr_growth)
mean_narci_selected_smallest_fwtofr_growth = statistics.mean(narci_selected_smallest_fwtofr_growth)
stdv_narci_selected_smallest_fwtofr_growth = statistics.stdev(narci_selected_smallest_fwtofr_growth)
mean_number_of_posts_selected_smallest_fwtofr_growth = statistics.mean(number_of_posts_selected_smallest_fwtofr_growth)
stdv_number_of_posts_selected_smallest_fwtofr_growth = statistics.stdev(number_of_posts_selected_smallest_fwtofr_growth)    
#------------------
ttt = heapq.nlargest(15, range(len(n_growth)), n_growth.__getitem__)  

fwtofr_selected_largest_n_growth = []
litofr_selected_largest_n_growth = []
narci_selected_largest_n_growth = []
number_of_posts_selected_largest_n_growth = []
name_largest_n_growth = []

for i in ttt:
    fwtofr_selected_largest_n_growth.append(ratio_of_followings_to_followers_for_all_users[i])
    litofr_selected_largest_n_growth.append(mean_like_to_follower_for_all_users[i])
    narci_selected_largest_n_growth.append(narcissism[i])
    number_of_posts_selected_largest_n_growth.append(number_of_posts[i])
    name_largest_n_growth.append(name[i+1]) #row 0 in names list belongs to labels so we add 1 to i

mean_fwtofr_selected_largest_n_growth = statistics.mean(fwtofr_selected_largest_n_growth)
stdv_fwtofr_selected_largest_n_growth = statistics.stdev(fwtofr_selected_largest_n_growth)
mean_litofr_selected_largest_n_growth = statistics.mean(litofr_selected_largest_n_growth)
stdv_litofr_selected_largest_n_growth = statistics.stdev(litofr_selected_largest_n_growth)
mean_narci_selected_largest_n_growth = statistics.mean(narci_selected_largest_n_growth)
stdv_narci_selected_largest_n_growth = statistics.stdev(narci_selected_largest_n_growth)
mean_number_of_posts_selected_largest_n_growth = statistics.mean (number_of_posts_selected_largest_n_growth)
stdv_number_of_posts_selected_largest_n_growth = statistics.stdev(number_of_posts_selected_largest_n_growth)
#------------------
ggg = heapq.nsmallest(15, range(len(n_growth)), n_growth.__getitem__)  

fwtofr_selected_smallest_n_growth = []
litofr_selected_smallest_n_growth = []
narci_selected_smallest_n_growth = []
number_of_posts_selected_smallest_n_growth = []
name_smallest_n_growth = []

for i in ggg:
    fwtofr_selected_smallest_n_growth.append(ratio_of_followings_to_followers_for_all_users[i])
    litofr_selected_smallest_n_growth.append(mean_like_to_follower_for_all_users[i])
    narci_selected_smallest_n_growth.append(narcissism[i])
    number_of_posts_selected_smallest_n_growth.append(number_of_posts[i])
    name_smallest_n_growth.append(name[i+1])#row 0 in names list belongs to labels so we add 1 to i

mean_fwtofr_selected_smallest_n_growth = statistics.mean(fwtofr_selected_smallest_n_growth)
stdv_fwtofr_selected_smallest_n_growth = statistics.stdev(fwtofr_selected_smallest_n_growth)
mean_litofr_selected_smallest_n_growth = statistics.mean(litofr_selected_smallest_n_growth)
stdv_litofr_selected_smallest_n_growth = statistics.stdev(litofr_selected_smallest_n_growth)
mean_narci_selected_smallest_n_growth = statistics.mean(narci_selected_smallest_n_growth)
stdv_narci_selected_smallest_n_growth = statistics.stdev(narci_selected_smallest_n_growth)
mean_number_of_posts_selected_smallest_n_growth = statistics.mean (number_of_posts_selected_smallest_n_growth)
stdv_number_of_posts_selected_smallest_n_growth = statistics.stdev(number_of_posts_selected_smallest_n_growth)



print (" --------------------------------------------- Statistics of both Datasets  ----------------------------------------- ")
print (" ------------------------------------------------ T.E.S.T D.A.T.A.S.E.T --------------------------------------------- ")
print (" mean_fwtofr : ", mean_fwtofr)
print (" stdv_fwtofr : ", stdv_fwtofr)
print (" mean_litofr : ", mean_litofr)
print (" stdv_litofr : ", stdv_litofr)
print (" mean_nar : ", mean_nar)
print (" stdv_nar : ", stdv_nar)
print (" mean_np : ", mean_np)
print (" stdv_np : ", stdv_np)
print (" --------------------------------------------------------------------------------------------------------------------- ")
print (" ------------------------------------ T.E.S.T D.A.T.A.S.E.T A.F.T.E.R 150 D.A.Y.S ------------------------------------ ")
print (" mean_fwtofr_150 : ", mean_fwtofr_150)
print (" stdv_fwtofr_150 : ", stdv_fwtofr_150)
print (" mean_litofr_150 : ", mean_litofr_150)
print (" stdv_litofr_150 : ", stdv_litofr_150)
print (" mean_nar_150 : ", mean_nar_150)
print (" stdv_nar_150 : ", stdv_nar_150)
print (" mean_np_150 : ", mean_np_150)
print (" stdv_np_150 : ", stdv_np_150)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" ---- P.E.A.R.S.O.N  C.O.E.F.I.C.I.E.N.T  O.F  'G.R.O.W.T.H  O.F  D.Y.N.A.M.I.C.S'  W.I.T.H  'E.A.C.H  O.T.H.E.R'  ---- ")
print (" n_growth_to_fwtofr_growth : ", n_growth_to_fwtofr_growth)
print (" n_growth_to_litofr_growth : ", n_growth_to_litofr_growth)
print (" fwtofr_growth_to_litofr_growth : ", fwtofr_growth_to_litofr_growth)
print (" np_growth_to_n_growth : ", np_growth_to_n_growth)
print (" np_growth_to_litofr_growth : ", np_growth_to_litofr_growth)
print (" np_growth_to_fwtofr_growth : ", np_growth_to_fwtofr_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -- S.P.E.A.R.M.A.N  C.O.E.F.I.C.I.E.N.T  O.F  'G.R.O.W.T.H  O.F  D.Y.N.A.M.I.C.S'  W.I.T.H  'E.A.C.H  O.T.H.E.R'  ---- ")
print (" n_growth_to_fwtofr_growth_s : ", n_growth_to_fwtofr_growth_s)
print (" n_growth_to_litofr_growth_s : ", n_growth_to_litofr_growth_s)
print (" fwtofr_growth_to_litofr_growth_s : ", fwtofr_growth_to_litofr_growth_s)
print (" np_growth_to_n_growth_s : ", np_growth_to_n_growth_s)
print (" np_growth_to_litofr_growth_s : ", np_growth_to_litofr_growth_s)
print (" np_growth_to_fwtofr_growth_s : ", np_growth_to_fwtofr_growth_s)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" ---- K.E.N.D.A.L.L  C.O.E.F.I.C.I.E.N.T  O.F  'G.R.O.W.T.H  O.F  D.Y.N.A.M.I.C.S'  W.I.T.H  'E.A.C.H  O.T.H.E.R'  ---- ")
print (" n_growth_to_fwtofr_growth_k : ", n_growth_to_fwtofr_growth_k)
print (" n_growth_to_litofr_growth_k : ", n_growth_to_litofr_growth_k)
print (" fwtofr_growth_to_litofr_growth_k : ", fwtofr_growth_to_litofr_growth_k)
print (" np_growth_to_n_growth_k : ", np_growth_to_n_growth_k)
print (" np_growth_to_litofr_growth_k : ", np_growth_to_litofr_growth_k)
print (" np_growth_to_fwtofr_growth_k : ", np_growth_to_fwtofr_growth_k)
print (" ----------------------------------------------------------------------------------------------------------------------- ")
print (" - P.E.A.R.S.O.N  C.O.E.F.I.C.I.E.N.T  O.F  'G.R.O.W.T.H  O.F  D.Y.N.A.M.I.C.S'  W.I.T.H  'E.A.R.L.Y  D.Y.N.A.M.I.C.S' - ")
print (" fwtofr_growth_to_fwtofr : ", fwtofr_growth_to_fwtofr)
print (" fwtofr_growth_to_litofr : ", fwtofr_growth_to_litofr)
print (" fwtofr_growth_to_narcissism : ", fwtofr_growth_to_narcissism)
print (" fwtofr_growth_to_number_of_posts : ", fwtofr_growth_to_number_of_posts)

print (" litofr_growth_to_fwtofr : ", litofr_growth_to_fwtofr)
print (" litofr_growth_to_litofr : ", litofr_growth_to_litofr)
print (" litofr_growth_to_narcissism : ", litofr_growth_to_narcissism)
print (" litofr_growth_to_number_of_posts : ", litofr_growth_to_number_of_posts)

print (" n_growth_to_fwtofr : ", n_growth_to_fwtofr_growth)
print (" n_growth_to_litofr : ", n_growth_to_litofr_growth)
print (" n_growth_to_narcissism : ", n_growth_to_narcissism)
print (" n_growth_to_number_of_posts : ", n_growth_to_number_of_posts)

print (" np_growth_to_fwtofr : ", np_growth_to_fwtofr)
print (" np_growth_to_litofr : ", np_growth_to_litofr)
print (" np_growth_to_narcissism : ", np_growth_to_narcissism)
print (" np_growth_to_number_of_posts : ", np_growth_to_number_of_posts)
print (" ----------------------------------------------------------------------------------------------------------------------- ")
print (" - S.P.E.A.R.M.A.N  C.O.E.F.I.C.I.E.N.T  O.F  'G.R.O.W.T.H  O.F  D.Y.N.A.M.I.C.S'  W.I.T.H  'E.A.R.L.Y  D.Y.N.A.M.I.C.S' - ")
print (" fwtofr_growth_to_fwtofr_s : ", fwtofr_growth_to_fwtofr_s)
print (" fwtofr_growth_to_litofr_s : ", fwtofr_growth_to_litofr_s)
print (" fwtofr_growth_to_narcissism_s : ", fwtofr_growth_to_narcissism_s)
print (" fwtofr_growth_to_number_of_posts_s : ", fwtofr_growth_to_number_of_posts_s)

print (" litofr_growth_to_fwtofr_s : ", litofr_growth_to_fwtofr_s)
print (" litofr_growth_to_litofr_s : ", litofr_growth_to_litofr_s)
print (" litofr_growth_to_narcissism_s : ", litofr_growth_to_narcissism_s)
print (" litofr_growth_to_number_of_posts_s : ", litofr_growth_to_number_of_posts_s)

print (" n_growth_to_fwtofr_s : ", n_growth_to_fwtofr_growth_s)
print (" n_growth_to_litofr_s : ", n_growth_to_litofr_growth_s)
print (" n_growth_to_narcissism_s : ", n_growth_to_narcissism_s)
print (" n_growth_to_number_of_posts_s : ", n_growth_to_number_of_posts_s)

print (" np_growth_to_fwtofr_s : ", np_growth_to_fwtofr_s)
print (" np_growth_to_litofr_s : ", np_growth_to_litofr_s)
print (" np_growth_to_narcissism_s : ", np_growth_to_narcissism_s)
print (" np_growth_to_number_of_posts_s : ", np_growth_to_number_of_posts_s)
print (" ----------------------------------------------------------------------------------------------------------------------- ")
print (" - K.E.N.D.A.L.L  C.O.E.F.I.C.I.E.N.T  O.F  'G.R.O.W.T.H  O.F  D.Y.N.A.M.I.C.S'  W.I.T.H  'E.A.R.L.Y  D.Y.N.A.M.I.C.S' - ")
print (" fwtofr_growth_to_fwtofr_k : ", fwtofr_growth_to_fwtofr_k)
print (" fwtofr_growth_to_litofr_k : ", fwtofr_growth_to_litofr_k)
print (" fwtofr_growth_to_narcissism_k : ", fwtofr_growth_to_narcissism_k)
print (" fwtofr_growth_to_number_of_posts_k : ", fwtofr_growth_to_number_of_posts_k)

print (" litofr_growth_to_fwtofr_k : ", litofr_growth_to_fwtofr_k)
print (" litofr_growth_to_litofr_k : ", litofr_growth_to_litofr_k)
print (" litofr_growth_to_narcissism_k : ", litofr_growth_to_narcissism_k)
print (" litofr_growth_to_number_of_posts_k : ", litofr_growth_to_number_of_posts_k)

print (" n_growth_to_fwtofr_k : ", n_growth_to_fwtofr_growth_k)
print (" n_growth_to_litofr_k : ", n_growth_to_litofr_growth_k)
print (" n_growth_to_narcissism_k : ", n_growth_to_narcissism_k)
print (" n_growth_to_number_of_posts_k : ", n_growth_to_number_of_posts_k)

print (" np_growth_to_fwtofr_k : ", np_growth_to_fwtofr_k)
print (" np_growth_to_litofr_k : ", np_growth_to_litofr_k)
print (" np_growth_to_narcissism_k : ", np_growth_to_narcissism_k)
print (" np_growth_to_number_of_posts_k : ", np_growth_to_number_of_posts_k)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -------------- S.T.A.T.I.S.T.I.C.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S    L.A.R.G.E.S.T   NP  G.R.O.W.T.H --------------- ")
print (" mean_fwtofr_selected_largest_np_growth : ", mean_fwtofr_selected_largest_np_growth)
print (" stdv_fwtofr_selected_largest_np_growth : ", stdv_fwtofr_selected_largest_np_growth)
print (" mean_litofr_selected_largest_np_growth : ", mean_litofr_selected_largest_np_growth)
print (" stdv_litofr_selected_largest_np_growth : ", stdv_litofr_selected_largest_np_growth)
print (" mean_narci_selected_largest_np_growth : ", mean_narci_selected_largest_np_growth)
print (" stdv_narci_selected_largest_np_growth : ", stdv_narci_selected_largest_np_growth)
print (" mean_number_of_posts_selected_largest_np_growth : ", mean_number_of_posts_selected_largest_np_growth)
print (" stdv_number_of_posts_selected_largest_np_growth : ", stdv_number_of_posts_selected_largest_np_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -------------- S.T.A.T.I.S.T.I.C.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S    S.M.A.L.L.E.S.T   NP  G.R.O.W.T.H --------- ")
print (" mean_fwtofr_selected_smallest_np_growth : ", mean_fwtofr_selected_smallest_np_growth)
print (" stdv_fwtofr_selected_smallest_np_growth : ", stdv_fwtofr_selected_smallest_np_growth)
print (" mean_litofr_selected_smallest_np_growth : ", mean_litofr_selected_smallest_np_growth)
print (" stdv_litofr_selected_smallest_np_growth : ", stdv_litofr_selected_smallest_np_growth)
print (" mean_narci_selected_smallest_np_growth : ", mean_narci_selected_smallest_np_growth)
print (" stdv_narci_selected_smallest_np_growth : ", stdv_narci_selected_smallest_np_growth)
print (" mean_number_of_posts_selected_smallest_np_growth : ", mean_number_of_posts_selected_smallest_np_growth)
print (" stdv_number_of_posts_selected_smallest_np_growth : ", stdv_number_of_posts_selected_smallest_np_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -------------- S.T.A.T.I.S.T.I.C.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S    L.A.R.G.E.S.T   FWtoFR  G.R.O.W.T.H ----------- ")
print (" mean_fwtofr_selected_largest_fwtofr_growth : ", mean_fwtofr_selected_largest_fwtofr_growth)
print (" stdv_fwtofr_selected_largest_fwtofr_growth : ", stdv_fwtofr_selected_largest_fwtofr_growth)
print (" mean_litofr_selected_largest_fwtofr_growth : ", mean_litofr_selected_largest_fwtofr_growth)
print (" stdv_litofr_selected_largest_fwtofr_growth : ", stdv_litofr_selected_largest_fwtofr_growth)
print (" mean_narci_selected_largest_fwtofr_growth : ", mean_narci_selected_largest_fwtofr_growth)
print (" stdv_narci_selected_largest_fwtofr_growth : ", stdv_narci_selected_largest_fwtofr_growth)
print (" mean_number_of_posts_selected_largest_fwtofr_growth : ", mean_number_of_posts_selected_largest_fwtofr_growth)
print (" stdv_number_of_posts_selected_largest_fwtofr_growth : ", stdv_number_of_posts_selected_largest_fwtofr_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -------------- S.T.A.T.I.S.T.I.C.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S    S.M.A.L.L.E.S.T  FWtoFR  G.R.O.W.T.H ---------- ")
print (" mean_fwtofr_selected_smallest_fwtofr_growth : ", mean_fwtofr_selected_smallest_fwtofr_growth)
print (" stdv_fwtofr_selected_smallest_fwtofr_growth : ", stdv_fwtofr_selected_smallest_fwtofr_growth)
print (" mean_litofr_selected_smallest_fwtofr_growth : ", mean_litofr_selected_smallest_fwtofr_growth)
print (" stdv_litofr_selected_smallest_fwtofr_growth : ", stdv_litofr_selected_smallest_fwtofr_growth)
print (" mean_narci_selected_smallest_fwtofr_growth : ", mean_narci_selected_smallest_fwtofr_growth)
print (" stdv_narci_selected_smallest_fwtofr_growth : ", stdv_narci_selected_smallest_fwtofr_growth)
print (" mean_number_of_posts_selected_smallest_fwtofr_growth : ", mean_number_of_posts_selected_smallest_fwtofr_growth)
print (" stdv_number_of_posts_selected_smallest_fwtofr_growth : ", stdv_number_of_posts_selected_smallest_fwtofr_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -------------- S.T.A.T.I.S.T.I.C.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S    L.A.R.G.E.S.T   LItoFR  G.R.O.W.T.H ----------- ")
print (" mean_fwtofr_selected_largest_litofr_growth : ", mean_fwtofr_selected_largest_litofr_growth)
print (" stdv_fwtofr_selected_largest_litofr_growth : ", stdv_fwtofr_selected_largest_litofr_growth)
print (" mean_litofr_selected_largest_litofr_growth : ", mean_litofr_selected_largest_litofr_growth)
print (" stdv_litofr_selected_largest_litofr_growth : ", stdv_litofr_selected_largest_litofr_growth)
print (" mean_narci_selected_largest_litofr_growth : ", mean_narci_selected_largest_litofr_growth)
print (" stdv_narci_selected_largest_litofr_growth : ", stdv_narci_selected_largest_litofr_growth)
print (" mean_number_of_posts_selected_largest_litofr_growth : ", mean_number_of_posts_selected_largest_litofr_growth)
print (" stdv_number_of_posts_selected_largest_litofr_growth : ", stdv_number_of_posts_selected_largest_litofr_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -------------- S.T.A.T.I.S.T.I.C.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S    S.M.A.L.L.E.S.T   LItoFR  G.R.O.W.T.H --------- ")
print (" mean_fwtofr_selected_smallest_litofr_growth : ", mean_fwtofr_selected_smallest_litofr_growth)
print (" stdv_fwtofr_selected_smallest_litofr_growth : ", stdv_fwtofr_selected_smallest_litofr_growth)
print (" mean_litofr_selected_smallest_litofr_growth : ", mean_litofr_selected_smallest_litofr_growth)
print (" stdv_litofr_selected_smallest_litofr_growth : ", stdv_litofr_selected_smallest_litofr_growth)
print (" mean_narci_selected_smallest_litofr_growth : ", mean_narci_selected_smallest_litofr_growth)
print (" stdv_narci_selected_smallest_litofr_growth : ", stdv_narci_selected_smallest_litofr_growth)
print (" mean_number_of_posts_selected_smallest_litofr_growth : ", mean_number_of_posts_selected_smallest_litofr_growth)
print (" stdv_number_of_posts_selected_smallest_litofr_growth : ", stdv_number_of_posts_selected_smallest_litofr_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -------------- S.T.A.T.I.S.T.I.C.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S    L.A.R.G.E.S.T   N  G.R.O.W.T.H ---------------- ")
print (" mean_fwtofr_selected_largest_n_growth : ", mean_fwtofr_selected_largest_n_growth)
print (" stdv_fwtofr_selected_largest_n_growth : ", stdv_fwtofr_selected_largest_n_growth)
print (" mean_litofr_selected_largest_n_growth : ", mean_litofr_selected_largest_n_growth)
print (" stdv_litofr_selected_largest_n_growth : ", stdv_litofr_selected_largest_n_growth)
print (" mean_narci_selected_largest_n_growth : ", mean_narci_selected_largest_n_growth)
print (" stdv_narci_selected_largest_n_growth : ", stdv_narci_selected_largest_n_growth)
print (" mean_number_of_posts_selected_largest_n_growth : ", mean_number_of_posts_selected_largest_n_growth)
print (" stdv_number_of_posts_selected_largest_n_growth : ", stdv_number_of_posts_selected_largest_n_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" -------------- S.T.A.T.I.S.T.I.C.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S    S.M.A.L.L.E.S.T   N  G.R.O.W.T.H -------------- ")
print (" mean_fwtofr_selected_smallest_n_growth : ", mean_fwtofr_selected_smallest_n_growth)
print (" stdv_fwtofr_selected_smallest_n_growth : ", stdv_fwtofr_selected_smallest_n_growth)
print (" mean_litofr_selected_smallest_n_growth : ", mean_litofr_selected_smallest_n_growth)
print (" stdv_litofr_selected_smallest_n_growth : ", stdv_litofr_selected_smallest_n_growth)
print (" mean_narci_selected_smallest_n_growth : ", mean_narci_selected_smallest_n_growth)
print (" stdv_narci_selected_smallest_n_growth : ", stdv_narci_selected_smallest_n_growth)
print (" mean_number_of_posts_selected_smallest_n_growth : ", mean_number_of_posts_selected_smallest_n_growth)
print (" stdv_number_of_posts_selected_smallest_n_growth : ", stdv_number_of_posts_selected_smallest_n_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" ------------------------------------- N.A.M.E.S  O.F  S.E.L.E.C.T.E.D  U.S.E.R.S   ----------------------------------- ")
print (" name_largest_np_growth : ", name_largest_np_growth)
print (" name_smallest_np_growth : ", name_smallest_np_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" name_largest_litofr_growth : ", name_largest_litofr_growth)
print (" name_smallest_litofr_growth : ", name_smallest_litofr_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" name_largest_fwtofr_growth : ", name_largest_fwtofr_growth)
print (" name_smallest_fwtofr_growth : ", name_smallest_fwtofr_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" name_largest_n_growth : ", name_largest_n_growth)
print (" name_smallest_n_growth : ", name_smallest_n_growth)
print (" ---------------------------------------------------------------------------------------------------------------------- ")
print (" name[0] : " , name[0])
print (" number_of_posts[0] : " , number_of_posts[0])
print (" number_of_followers[0] : " , number_of_followers[0])
print (" number_of_followings[0] : " , number_of_followings[0])
print (" mean_like_for_all_users[0] : " , mean_like_for_all_users[0])
print (" number_of_self_picture_posts_form_9_previous_posts[0] : " , number_of_self_picture_posts_form_9_previous_posts[0])
print (" narcissism[0] : " , narcissism[0])
print (" ratio_of_followings_to_followers_for_all_users[0] : " , ratio_of_followings_to_followers_for_all_users[0])
print (" mean_like_to_follower_for_all_users[0] : " , mean_like_to_follower_for_all_users[0])
print (" fwtofr_growth[0] : " , fwtofr_growth[0])
##import matplotlib.pyplot as plt
##plt.plot (fwtofr_growth,ratio_of_followings_to_followers_for_all_users,"o")
##plt.show()
##plt.plot (litofr_growth,mean_like_to_follower_for_all_users,"o")
##plt.show()
##plt.plot (np_growth,number_of_posts,"o")
##plt.show()


input("\n press enter key to exit.")
